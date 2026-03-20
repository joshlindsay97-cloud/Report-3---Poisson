#!/usr/bin/env python3
"""
Convert CSV outputs to XLSX for easy viewing.

Run:
  python to_xlsx.py --in out/task5_comparison.csv --out out/task5_comparison.xlsx
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--in", dest="in_csv", type=str, required=True)
    p.add_argument("--out", dest="out_xlsx", type=str, required=True)
    return p.parse_args()


def main() -> None:
    a = parse_args()
    in_path = Path(a.in_csv)
    out_path = Path(a.out_xlsx)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "data"

    with in_path.open(newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # autosize columns (rough)
    for c_idx in range(1, len(rows[0]) + 1):
        col = get_column_letter(c_idx)
        max_len = max(len(str(rows[r][c_idx - 1])) if c_idx - 1 < len(rows[r]) else 0 for r in range(min(len(rows), 500)))
        ws.column_dimensions[col].width = min(60, max(10, max_len + 2))

    wb.save(out_path)
    print(f"wrote {out_path}")


if __name__ == "__main__":
    main()
